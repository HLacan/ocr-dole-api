#!/usr/bin/env python3
"""
extract_dole.py  —  OCR Dole: Excel de control semanal de embarques

Relaciones clave:
  1 BL (GYEPRQ...) → N facturas (por contenedor compartido)
  1 Factura        → N fitos (por contenedor + bultos que deben cuadrar)
  Todo se cruza por CONTENEDOR ([A-Z]{4}[0-9]{7}) — NO por nombre de carpeta.

Formato de entrada:
  --input-folder: TODOS los archivos de la semana (facturas, fitos,
    manifiestos). No importa si vienen sueltos o en subcarpetas — el
    script recorre todo recursivamente y cruza cada documento con su
    contenedor según el propio contenido del PDF.
  --bl-file: el PDF combinado de BLs (una página = un BL, ej. "BLS DAC608
    VIA PQU ORIG.pdf", ~100 páginas). Si se omite, el script busca
    automáticamente dentro de --input-folder un archivo con nombre de BL
    (ej. que empiece con "BLS").

Uso:
  python3 extract_dole.py \
    --input-folder  "ruta/DOLE SEMANA 6" \
    --bl-file       "ruta/BLS DAC608 VIA PQU ORIG.pdf" \
    --output        "ruta/DOLE-2026-N.xlsx" \
    --semana        "6" \
    [--marchamo-file "ruta/LISTADO DE MARCHAMOS.xlsx"]
"""

import os, re, sys, argparse, subprocess

def ensure_deps():
    try:
        import pdfplumber, openpyxl, fitz, pytesseract
        from PIL import Image
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install",
                        "pdfplumber", "openpyxl", "pymupdf", "pytesseract", "Pillow",
                        "--break-system-packages", "-q"], check=True)
ensure_deps()

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONTAINER_RE = re.compile(r'\b([A-Z]{4}[0-9]{7})\b')

# ─────────────────────────────────────────────
# Extracción de texto — con OCR de respaldo
# ─────────────────────────────────────────────
# En la práctica, algunos proveedores (ej. UBESA) generan PDFs con una
# fuente embebida sin mapa Unicode real: pdfplumber/pymupdf extraen texto
# corrupto (caracteres de control). Se detecta ese caso y se usa OCR sobre
# la imagen renderizada de la página como respaldo.

def _is_garbled(text):
    if not text or len(text.strip()) < 20:
        return True
    # pdfplumber representa glifos sin mapa Unicode como "(cid:123)" — es un
    # indicador inequívoco de fuente rota, independiente de si el resto del
    # texto "parece" imprimible.
    if text.count('(cid:') >= 3:
        return True
    sample = text[:2000]
    printable = sum(1 for c in sample if c.isprintable() and
                     (c.isalnum() or c in " .,:;/-()°ºÁÉÍÓÚáéíóúÑñ%$#"))
    return printable / max(len(sample), 1) < 0.6


def _ocr_pdf(path):
    try:
        import fitz
        import pytesseract
        from PIL import Image
        import io
        import gc

        doc = fitz.open(path)
        parts = []
        for page in doc:
            # 150 dpi en vez de 200: sigue siendo suficiente para que
            # tesseract lea bien estos certificados, pero usa bastante
            # menos memoria por pagina (la relacion es cuadratica con el
            # dpi) -- importante en hostings con poca RAM (ej. 512MB).
            pix = page.get_pixmap(dpi=150)
            png_bytes = pix.tobytes("png")
            pix = None  # liberar el pixmap nativo cuanto antes

            img = Image.open(io.BytesIO(png_bytes))
            try:
                parts.append(pytesseract.image_to_string(img, lang="eng+spa"))
            except Exception:
                parts.append(pytesseract.image_to_string(img, lang="eng"))
            finally:
                img.close()
                del img, png_bytes
                gc.collect()  # evita que la memoria de esta pagina se acumule sobre la siguiente

        doc.close()
        return "\n".join(parts)
    except Exception:
        return ""


def pdf_text(path):
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception:
        text = ""
    if _is_garbled(text):
        ocr_text = _ocr_pdf(path)
        if ocr_text.strip():
            return ocr_text
    return text


def xlsx_text(path):
    """Convierte todas las celdas de un xlsx/xlsm a texto plano, para poder
    aplicarle las mismas búsquedas por palabra clave/regex que a un PDF."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return ""
    lines = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    lines.append(str(cell))
    return "\n".join(lines)


def xls_legacy_text(path):
    """Convierte un .xls VIEJO (formato binario, no XML) a texto plano.
    openpyxl NO puede leer este formato — en la práctica, las facturas de
    UBESA llegan exactamente así (ej. 'ECFPOFACTURA QUETZAL - DAC609-
    INV6750.xls'), y con openpyxl ese archivo se leería como texto vacío,
    perdiendo silenciosamente la factura completa. Se usa xlrd en su lugar."""
    try:
        import xlrd
    except ImportError:
        return ""
    try:
        wb = xlrd.open_workbook(path)
    except Exception:
        return ""
    lines = []
    for sheet_idx in range(wb.nsheets):
        sh = wb.sheet_by_index(sheet_idx)
        for r in range(sh.nrows):
            for cell in sh.row_values(r):
                if cell not in (None, ""):
                    lines.append(str(cell))
    return "\n".join(lines)


def doc_text(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        return xls_legacy_text(path)
    if ext in ('.xlsx', '.xlsm'):
        return xlsx_text(path)
    return pdf_text(path)

# ─────────────────────────────────────────────
# Clasificación de archivos
# ─────────────────────────────────────────────
# Solo BL y MANIFIESTO se clasifican por NOMBRE (es confiable). Facturas y
# fitos se clasifican por CONTENIDO en classify_and_extract_all(), porque en
# la práctica los nombres reales no siguen ningún patrón fijo (ej. no todas
# las facturas UBESA se llaman "EC-UB-*", ni todos los fitos dicen "PHYTO").

def classify_file(filename):
    name = filename.upper()
    if 'CONSULTAMANIFIESTO' in name or ('CONSULTA' in name and 'MANIFIESTO' in name):
        return 'MANIFIESTO'
    if name.startswith('BLS') or (name.startswith('BL') and 'DPC' in name):
        return 'BL'
    return 'UNKNOWN'

def file_mtime(path):
    try: return os.path.getmtime(path)
    except: return 0

def newest_file(paths):
    return max(paths, key=file_mtime) if paths else None

# ─────────────────────────────────────────────
# Extracción BL — PDF combinado (1 página = 1 BL)
# ─────────────────────────────────────────────

def split_bl_pdf(bl_pdf_path):
    """
    El BL llega como un solo PDF de varias páginas (ej. 100 páginas).
    Cada página es un BL individual, identificado por su código
    GYEPRQ###### presente en el texto de esa página.

    Retorna: dict { "GYEPRQ356004": texto_pagina, ... }
    Si el mismo código aparece en más de una página (poco común), se
    concatena el texto de todas esas páginas.
    """
    pages_by_code = {}
    try:
        with pdfplumber.open(bl_pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                m = re.search(r'\b(GYEPRQ\d+)\b', text, re.IGNORECASE)
                if not m:
                    continue
                code = m.group(1).upper()
                if code in pages_by_code:
                    pages_by_code[code] += "\n" + text
                else:
                    pages_by_code[code] = text
    except Exception as e:
        print(f"  ERROR leyendo BL combinado {bl_pdf_path}: {e}")
    return pages_by_code


def extract_bl_from_text(text, bl_code):
    """
    Extrae de una página del BL combinado:
      BL, CONTENEDOR, DESTINO (Place of delivery), TIPO (ISO container type)

    Estructura clave:
      Línea "GUAYAQUIL {PLACE_OF_DELIVERY}"        → DESTINO
      Línea "1 x {TYPE} CONTAINERS"                → TIPO (texto libre)
      Tabla "XXXX1234567 SEAL 40 RF 960 BOX ..."   → CONTENEDOR + TIPO (tabla)

    El TIPO de texto libre y el de la tabla a veces no coinciden (ej. RC
    vs RF). En ese caso se usa el de la TABLA (específico del contenedor
    real) y se genera una advertencia para revisión manual.
    """
    lines = text.split('\n')

    bl = bl_code or "???"
    contenedor = "???"
    destino = "???"
    tipo_texto = None
    tipo_tabla = None
    warning = None

    # DESTINO: línea "GUAYAQUIL {Place of Delivery}" que contiene el país destino
    for line in lines:
        stripped = line.strip()
        if stripped.upper().startswith("GUAYAQUIL ") and "USA" in stripped.upper():
            destino_raw = stripped[len('GUAYAQUIL '):].strip()
            if destino_raw and not re.match(r'^\d', destino_raw):
                destino = destino_raw
                break

    # TIPO (texto libre): "1 x 40RV CONTAINERS"
    m_tipo_texto = re.search(r'\d+\s*x\s*(40[A-Z]{2})\s+CONTAINER', text, re.IGNORECASE)
    if m_tipo_texto:
        tipo_texto = m_tipo_texto.group(1).upper()

    # CONTENEDOR + TIPO (tabla): "CAAU4102550 132547E 40 RF 960 BOX 18941.76 KGS"
    m_row = re.search(
        r'\b([A-Z]{4}[0-9]{7})\s+\S+\s+(\d{2})\s*([A-Z]{2})\s+(\d+)\s*BOX',
        text, re.IGNORECASE
    )
    if m_row:
        contenedor = m_row.group(1).upper()
        tipo_tabla = (m_row.group(2) + m_row.group(3)).upper()
    else:
        containers = [c for c in CONTAINER_RE.findall(text)
                      if not c.startswith('GYE') and not c.startswith('AG1')]
        if containers:
            contenedor = containers[0]

    tipo = tipo_tabla or tipo_texto or "???"
    if tipo_texto and tipo_tabla and tipo_texto != tipo_tabla:
        warning = (f"TIPO no coincide en {bl} [{contenedor}]: "
                   f"texto libre={tipo_texto} vs tabla={tipo_tabla} "
                   f"→ se usó el de la tabla ({tipo_tabla}), revisar manualmente")

    return {
        "BL": bl,
        "CONTENEDOR": contenedor,
        "DESTINO": destino,
        "TIPO": tipo,
    }, warning

# ─────────────────────────────────────────────
# Extracción MANIFIESTO
# ─────────────────────────────────────────────

def extract_manifiesto(path):
    text = pdf_text(path)
    m = re.search(r'Numero de manifiesto\s+(AG\d+)', text)
    if m: return m.group(1)
    m = re.search(r'(AG\d{9,})', text)
    return m.group(1) if m else "???"

# ─────────────────────────────────────────────
# Extracción FITO
# ─────────────────────────────────────────────

def _extract_containers_list(text):
    """
    Busca una línea tipo 'CONTAINERS: DFIU1234567-DFIU7654321-...' (varios,
    típico de facturas/fitos de "otros exportadores" que agrupan varios
    contenedores) o 'CONTAINER: DFIU1234567' (uno solo, típico de UBESA).
    Si no hay línea explícita, cae a buscar cualquier código de contenedor
    presente en todo el documento.
    """
    compact = re.sub(r'\s+', '', text.upper())
    m = re.search(r'CONTAINERS?:([A-Z0-9\-,]+)', compact)
    if m:
        codes = re.findall(r'[A-Z]{4}[0-9]{7}', m.group(1))
        if codes:
            return codes
    return [c for c in CONTAINER_RE.findall(text) if not c.startswith('GYE')]


def _extract_bultos_total(text):
    upper = text.upper()
    # Prioridad 1: facturas en xlsx suelen traer una celda "Cantidad Total: N"
    m = re.search(r'CANTIDAD\s*TOTAL\s*:?\s*(\d{2,6})', upper)
    if m:
        return int(m.group(1))
    # Prioridad 2: "4800 Box" / "4800 Boxes" (fitos y facturas PDF)
    m = re.search(r'(\d{2,6})\s*BOX(?:ES)?\b', text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    # Prioridad 3: "4800 CAJAS" (plural exacto — "CAJA" singular suele ser
    # parte de un nombre de producto/código, no una cantidad, y daría falsos
    # positivos como "6100 CAJA DOLE CONVENCIONAL")
    m = re.search(r'(\d{2,6})\s*CAJAS\b', upper)
    if m:
        return int(m.group(1))
    return None


def _extract_proveedor_from_filename(filename):
    """Toma las primeras palabras del nombre de archivo, antes de llegar a
    un codigo de puerto (FPO/GPT/ILG) o una palabra clave tipo FACT/PHYTO,
    como mejor intento de proveedor. Es el fallback mas confiable cuando NO
    hay estructura de carpetas disponible (ej. estamos procesando archivos
    ya aplanados por /process-email, donde se perdio el path original)."""
    nombre = os.path.splitext(os.path.basename(filename))[0]
    palabras = nombre.replace('#', ' ').split()
    detener_en = {'FPO', 'GPT', 'ILG', 'FACT', 'FACTURA', 'PHYTO', 'ORG', 'BANANAS', 'BB', 'ACT'}
    resultado = []
    for palabra in palabras:
        if palabra.upper() in detener_en or re.match(r'^\d', palabra):
            break
        resultado.append(palabra)
    texto = ' '.join(resultado).upper().strip()
    return texto if texto else "???"


def _extract_proveedor_from_path(path, text=""):
    """El proveedor se detecta en este orden (de mas a menos confiable):
    1) CONTENIDO del documento (ej. "UNION DE BANANEROS ECUATORIANOS" -> UBESA)
    2) La carpeta contenedora, SI existe una carpeta real de exportador
       (ej. .../OTROS EXPORTADORES/LUDERSON/archivo.pdf -> LUDERSON)
    3) El propio NOMBRE DE ARCHIVO -- necesario cuando el archivo ya viene
       suelto, sin ninguna carpeta que lo identifique (ej. procesado via
       /process-email, donde ya no existe el path original)."""
    upper_text = text.upper()
    if 'UNION DE BANANEROS ECUATORIANOS' in upper_text or 'UBESA' in upper_text:
        return 'UBESA'

    parts = [p.upper() for p in os.path.normpath(path).split(os.sep)]
    if 'UBESA' in parts:
        return 'UBESA'
    fname_upper = os.path.basename(path).upper()
    if fname_upper.startswith('EC') and 'FACTURA' in fname_upper:
        return 'UBESA'

    parent = os.path.basename(os.path.dirname(path)).upper().strip()
    carpetas_no_validas = ('OTROS EXPORTADORES', '', 'TMP') 
    es_carpeta_temporal = parent.startswith('OCR_DOLE_BATCH') or parent.startswith('/TMP')
    if parent and parent not in carpetas_no_validas and not es_carpeta_temporal:
        return parent

    return _extract_proveedor_from_filename(path)


def extract_fito(path):
    """Returns {number, bultos, containers, item}"""
    text = doc_text(path)
    upper = text.upper()
    compact = re.sub(r'\s+', '', text)

    m = re.search(r'N[°ºo]\.?([0-9]{15,30}P)', compact, re.IGNORECASE)
    number = m.group(1).upper() if m else None
    if not number:
        m2 = re.search(r'([0-9]{15,30}P)\b', compact)
        number = m2.group(1).upper() if m2 else "???"

    bultos = _extract_bultos_total(text)
    containers = _extract_containers_list(text)
    item = "PLATANO" if ("PLATANO" in upper or "PLANTAIN" in upper
                         or "PLANT" in os.path.basename(path).upper()) else "BANANO"
    proveedor = _extract_proveedor_from_path(path, text)

    return {"number": number, "bultos": bultos, "containers": containers, "item": item, "proveedor": proveedor}

# ─────────────────────────────────────────────
# Extracción FACTURA
# ─────────────────────────────────────────────

def extract_factura(path):
    """Returns {proveedor, factura, item, containers, bultos}"""
    text = doc_text(path)
    upper = text.upper()

    proveedor = _extract_proveedor_from_path(path, text)

    # Número de factura: formato estándar Ecuador (001-002-000000123456),
    # luego "FACT #1234" / "FACT 1234", luego "No. 123456789..." genérico
    factura = "???"
    m = re.search(r'(\d{3}-\d{3}-\d{6,9})', text)
    if m:
        factura = m.group(1)
    else:
        m = re.search(r'FACT(?:URA)?\.?\s*#?\s*(\d{3,10})', upper)
        if m:
            factura = m.group(1)
        else:
            m = re.search(r'\bNo\.?\s*[:\-]?\s*(\d{9,15})\b', text)
            if m:
                factura = m.group(1)

    item = "PLATANO" if ("PLATANO" in upper or "PLANTAIN" in upper
                         or "PLANT" in os.path.basename(path).upper()) else "BANANO"

    containers = _extract_containers_list(text)
    bultos = _extract_bultos_total(text)

    return {
        "proveedor": proveedor, "factura": factura, "item": item,
        "containers": containers, "bultos": bultos,
    }

# ─────────────────────────────────────────────
# Recorrer TODOS los archivos (sin importar estructura de carpetas)
# ─────────────────────────────────────────────

def classify_and_extract_all(input_folder, exclude_path=None):
    """
    Recorre input_folder recursivamente (sirve tanto si los archivos vienen
    sueltos como organizados en subcarpetas) y clasifica/extrae cada
    factura/fito/manifiesto **por contenido** (los nombres reales no siguen
    un patrón fijo). Cada archivo se lee una sola vez.
    """
    exclude_abs = os.path.abspath(exclude_path) if exclude_path else None
    facturas, fitos, manifiestos = [], [], []
    for root, _dirs, files in os.walk(input_folder):
        for fname in files:
            fpath = os.path.join(root, fname)
            if exclude_abs and os.path.abspath(fpath) == exclude_abs:
                continue

            name_upper = fname.upper()
            ext = os.path.splitext(fname)[1].lower()

            # BL combinado: se procesa aparte, se ignora aquí
            if name_upper.startswith('BLS') or (name_upper.startswith('BL') and 'DPC' in name_upper):
                continue
            # Manifiesto: confiable por nombre
            if 'CONSULTAMANIFIESTO' in name_upper or ('CONSULTA' in name_upper and 'MANIFIESTO' in name_upper):
                manifiestos.append({"number": extract_manifiesto(fpath), "_mtime": file_mtime(fpath)})
                continue

            if ext not in ('.pdf', '.xlsx', '.xlsm', '.xls'):
                continue

            text = doc_text(fpath)
            upper = text.upper()

            if 'FITOSANITARIO' in upper or 'PHYTOSANITARY' in upper:
                fi = extract_fito(fpath)
                fi['_mtime'] = file_mtime(fpath)
                fitos.append(fi)
            elif 'FACTURA' in upper or 'INVOICE' in upper or 'PROFORMA' in upper:
                fd = extract_factura(fpath)
                fd['_mtime'] = file_mtime(fpath)
                facturas.append(fd)
            elif 'MANIFIESTO' in upper:
                manifiestos.append({"number": extract_manifiesto(fpath), "_mtime": file_mtime(fpath)})
            # si no matchea nada: se ignora (UNKNOWN)
    return facturas, fitos, manifiestos


def dedup_by_key(items, key):
    """Dedup por valor de `key`, gana el más reciente (_mtime). Los items
    sin valor válido de `key` (None o '???') NO se colapsan entre sí — se
    conservan todos, porque agruparlos perdería datos distintos."""
    best = {}
    passthrough = []
    for it in items:
        k = it.get(key)
        if not k or k == "???":
            passthrough.append(it)
            continue
        mt = it.get('_mtime', 0)
        if k not in best or mt > best[k][1]:
            best[k] = (it, mt)
    return [v[0] for v in best.values()] + passthrough


# ─────────────────────────────────────────────
# Construir fila del Excel para un CONTENEDOR
# ─────────────────────────────────────────────

def build_row(container, bl_data, bl_warning, facturas, fitos, manifiestos):
    row = {
        "BL": bl_data["BL"] if bl_data else "???",
        "PROVEEDOR": "???", "CONTENEDOR": container,
        "TIPO": bl_data["TIPO"] if bl_data else "???",
        "DESTINO": bl_data["DESTINO"] if bl_data else "???",
        "FACTURA": "???", "ITEM": "???", "FITO": "???", "MANIFIESTO": "???",
    }
    warnings = []
    if bl_warning:
        warnings.append(bl_warning)
    if not bl_data:
        warnings.append(f"{container}: no se encontró BL para este contenedor")

    # FACTURAS cuyo listado de contenedores incluye este contenedor
    # (una factura puede cubrir varios contenedores a la vez)
    matched_facturas = [f for f in facturas if container in f.get("containers", [])]
    if matched_facturas:
        row["PROVEEDOR"] = matched_facturas[0]["proveedor"]
        row["ITEM"]      = matched_facturas[0]["item"]
        nums = [f["factura"] for f in matched_facturas if f["factura"] != "???"]
        row["FACTURA"] = " / ".join(dict.fromkeys(nums)) if nums else "???"
    else:
        warnings.append(f"{container}: no se encontró factura para este contenedor")

    # FITOS cuyo listado de contenedores incluye este contenedor
    matched_fitos = [fi for fi in fitos if container in fi.get("containers", [])]
    if matched_fitos:
        row["FITO"] = " / ".join(dict.fromkeys(fi["number"] for fi in matched_fitos))
        if row["ITEM"] == "???" and matched_fitos[0].get("item"):
            row["ITEM"] = matched_fitos[0]["item"]
    else:
        warnings.append(f"{container}: no se encontró fito para este contenedor")

    # MANIFIESTO — normalmente uno solo cubre todo el embarque/semana
    if len(manifiestos) == 1:
        row["MANIFIESTO"] = manifiestos[0]["number"]
    elif len(manifiestos) > 1:
        row["MANIFIESTO"] = manifiestos[0]["number"]
        warnings.append(f"{container}: hay {len(manifiestos)} manifiestos distintos, "
                         f"se usó el primero — revisar manualmente")

    # Validación de bultos: el FITO es la referencia (una factura puede
    # cubrir varios contenedores con un solo total combinado, así que solo
    # se compara cuando hay exactamente 1 factura y 1 fito para este
    # contenedor y ambos traen una cifra.
    if len(matched_facturas) == 1 and len(matched_fitos) == 1:
        bf  = matched_facturas[0].get("bultos")
        bfi = matched_fitos[0].get("bultos")
        if bf and bfi and bf != bfi:
            warnings.append(
                f"BULTOS NO CUADRAN en {container}: Factura={bf} Fito={bfi} (el fito es la referencia)"
            )

    return row, warnings

# ─────────────────────────────────────────────
# LISTADO DE MARCHAMOS
# ─────────────────────────────────────────────

def split_bl_pdf_pages(pdf_bytes, generar_pdfs=True):
    """
    Version 'API' de split_bl_pdf(): en vez de leer de una ruta y devolver
    solo texto por código, recibe los BYTES del PDF combinado y devuelve,
    por cada página, el PDF de una sola página (bytes) + los datos ya
    extraídos con extract_bl_from_text (misma lógica ya validada).

    generar_pdfs=False se salta por completo la generacion de cada PDF
    individual (PdfWriter + buffer) -- ahorra memoria cuando el llamador
    de todos modos no va a usar esos bytes (ej. incluir_archivos=false).

    Retorna: list of {"gyeprq", "contenedor", "destino", "tipo",
                       "warning", "pdf_bytes"} (pdf_bytes=None si
                       generar_pdfs=False)
    """
    import io
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(io.BytesIO(pdf_bytes)) if generar_pdfs else None
    resultados = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            m = re.search(r'\b(GYEPRQ\d+)\b', text, re.IGNORECASE)
            code = m.group(1).upper() if m else None

            bl_data, warning = extract_bl_from_text(text, code)

            pdf_bytes_pagina = None
            if generar_pdfs:
                writer = PdfWriter()
                writer.add_page(reader.pages[i])
                buf = io.BytesIO()
                writer.write(buf)
                pdf_bytes_pagina = buf.getvalue()

            resultados.append({
                "gyeprq": bl_data["BL"],
                "contenedor": bl_data["CONTENEDOR"],
                "destino": bl_data["DESTINO"],
                "tipo": bl_data["TIPO"],
                "pagina": i + 1,
                "warning": warning,
                "pdf_bytes": pdf_bytes_pagina,
            })
    return resultados


def classify_single_document(path):
    """
    Clasifica y extrae UN documento suelto (factura, fito, manifiesto o
    marchamos) sin necesitar el resto de la semana — pensado para el
    endpoint /classify-document, que se llama una vez por archivo a medida
    que van llegando en cada correo.

    Retorna: {"tipo": "FACTURA"|"FITO"|"MANIFIESTO"|"MARCHAMOS"|"DESCONOCIDO",
              ...campos propios de ese tipo...}
    """
    fname_upper = os.path.basename(path).upper()
    ext = os.path.splitext(path)[1].lower()

    if 'CONSULTAMANIFIESTO' in fname_upper or ('CONSULTA' in fname_upper and 'MANIFIESTO' in fname_upper):
        numero = extract_manifiesto(path)
        text = pdf_text(path)
        codigos = sorted(set(re.findall(r'GYEPRQ\d+', text)))
        return {"tipo": "MANIFIESTO", "numero": numero, "gyeprq_cubiertos": codigos}

    if ext in ('.xlsx', '.xlsm') and 'MARCHAMO' in fname_upper:
        lookup = load_marchamo_lookup(path)
        return {"tipo": "MARCHAMOS", "contenedor_a_marchamo": lookup}

    if ext not in ('.pdf', '.xlsx', '.xlsm', '.xls'):
        return {"tipo": "DESCONOCIDO", "motivo": f"extensión no soportada: {ext}"}

    text = doc_text(path)
    upper = text.upper()

    if 'FITOSANITARIO' in upper or 'PHYTOSANITARY' in upper:
        fi = extract_fito(path)
        return {"tipo": "FITO", **fi}

    if 'FACTURA' in upper or 'INVOICE' in upper or 'PROFORMA' in upper:
        fd = extract_factura(path)
        return {"tipo": "FACTURA", **fd}

    if 'MANIFIESTO' in upper:
        numero = extract_manifiesto(path)
        codigos = sorted(set(re.findall(r'GYEPRQ\d+', text)))
        return {"tipo": "MANIFIESTO", "numero": numero, "gyeprq_cubiertos": codigos}

    return {"tipo": "DESCONOCIDO", "motivo": "no se encontraron palabras clave (FACTURA/FITOSANITARIO/MANIFIESTO)"}


EXTENSIONES_VALIDAS = ('.pdf', '.xls', '.xlsx', '.xlsm')


def flatten_zip_bytes(filename, content):
    """
    Aplana recursivamente un archivo que puede ser un zip (con zips
    anidados adentro, a cualquier profundidad) usando la libreria estandar
    'zipfile' -- no hay restricciones de modulos aqui como en n8n.

    Si 'filename' no es un zip, simplemente lo devuelve tal cual (siempre
    que su extension nos interese). Descarta automaticamente las entradas
    de carpeta y cualquier archivo con extension que no nos sirva.

    Retorna: list of {"filename": str, "content": bytes}
    """
    import zipfile
    import io

    resultados = []

    def _flatten(nombre, datos):
        if nombre.lower().endswith('.zip'):
            with zipfile.ZipFile(io.BytesIO(datos)) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    nombre_interno = info.filename.split('/')[-1]
                    if not nombre_interno:
                        continue
                    _flatten(nombre_interno, zf.read(info))
        else:
            ext = os.path.splitext(nombre)[1].lower()
            if ext in EXTENSIONES_VALIDAS:
                resultados.append({"filename": nombre, "content": datos})

    _flatten(filename, content)
    return resultados


def load_marchamo_lookup(marchamo_file):
    """CONTENEDOR → 'SAT-GT-{numero}'"""
    lookup = {}
    if not marchamo_file or not os.path.isfile(marchamo_file):
        return lookup
    try:
        wb = openpyxl.load_workbook(marchamo_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=8, values_only=True):
            if row[1] and row[5]:
                container    = str(row[1]).strip()
                marchamo_num = str(int(row[5])) if isinstance(row[5], float) else str(row[5]).strip()
                lookup[container] = f"SAT-GT-{marchamo_num}"
    except Exception as e:
        print(f"  ADVERTENCIA: No se pudo leer LISTADO DE MARCHAMOS: {e}")
    return lookup

# ─────────────────────────────────────────────
# Generación del Excel
# ─────────────────────────────────────────────

HEADERS = [
    "No.", "FILE", "PROVEEDOR", "BL", "CONTENEDOR", "TIPO", "DESTINO",
    "FACTURA", "ITEM", "FITO", "MANIFIESTO",
    "PILOTO", "LICENCIA", "PLACA", "FIANZA",
    "CODIGO DE TT", "TRANSPORTE ", "SELECTIVO", "DUCA",
    "RETENCION", "TRANSITO", "MARCHAMO"
]

# Siempre manuales (amarillo)
MANUAL_COLS = {
    "FILE", "PILOTO", "LICENCIA", "PLACA", "FIANZA",
    "CODIGO DE TT", "TRANSPORTE ", "SELECTIVO", "DUCA",
    "RETENCION", "TRANSITO",
}

YELLOW_FILL  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL  = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
AUTO_FILL    = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
MARCH_FILL   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DEST_FILL    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # naranja claro

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(size=9)
thin         = Side(style='thin', color='BBBBBB')
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

COL_WIDTHS = {
    1: 5,   # No.
    2: 8,   # FILE
    3: 14,  # PROVEEDOR
    4: 16,  # BL
    5: 14,  # CONTENEDOR
    6: 7,   # TIPO
    7: 22,  # DESTINO
    8: 28,  # FACTURA
    9: 9,   # ITEM
    10: 36, # FITO
    11: 14, # MANIFIESTO
    12: 14, # PILOTO
    13: 12, # LICENCIA
    14: 10, # PLACA
    15: 10, # FIANZA
    16: 12, # CODIGO DE TT
    17: 12, # TRANSPORTE
    18: 10, # SELECTIVO
    19: 10, # DUCA
    20: 10, # RETENCION
    21: 10, # TRANSITO
    22: 16, # MARCHAMO
}


def generate_excel(rows, output_path, semana, marchamo_lookup=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    # Excel prohíbe : \ / ? * [ ] en el nombre de hoja, y limita a 31 caracteres
    safe_title = re.sub(r'[:\\/?*\[\]]', '-', f"SEMANA {semana}")[:31] or "SEMANA"
    ws.title = safe_title

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    tc = ws.cell(row=1, column=1, value=f"SEMANA {semana}")
    tc.font = Font(bold=True, size=14)
    tc.alignment = Alignment(horizontal='center')

    # Encabezados (fila 3)
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    # Datos
    for ri, row_data in enumerate(rows, 4):
        cont = row_data.get("CONTENEDOR", "")
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            cell.font   = DATA_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            if h == "No.":
                cell.value = ri - 3

            elif h in MANUAL_COLS:
                cell.value = None
                cell.fill  = YELLOW_FILL

            elif h == "MARCHAMO":
                if marchamo_lookup and cont in marchamo_lookup:
                    cell.value = marchamo_lookup[cont]
                    cell.fill  = MARCH_FILL
                else:
                    cell.value = None
                    cell.fill  = YELLOW_FILL

            elif h == "DESTINO":
                # Auto-llenado pero resaltado en naranja claro para que sea visible
                # como campo clave para asignación de pilotos
                val = row_data.get("DESTINO", "")
                cell.value = val if val != "???" else ""
                cell.fill  = DEST_FILL if val and val != "???" else MISSING_FILL

            else:
                val = row_data.get(h.strip(), "")
                if val == "???":
                    cell.value = ""
                    cell.fill  = MISSING_FILL
                else:
                    cell.value = val
                    cell.fill  = AUTO_FILL

    # Nota al pie
    fr = len(rows) + 5
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=len(HEADERS))
    nc = ws.cell(row=fr, column=1,
                 value="Azul=auto | Naranja=DESTINO (clave para asignación de piloto) | Amarillo=captura manual | Verde=marchamo auto | Rosado=no encontrado")
    nc.font = Font(italic=True, size=8, color="666666")

    # Anchos y alturas
    for ci, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    for r in range(4, len(rows) + 4):
        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "C4"
    wb.save(output_path)

# ─────────────────────────────────────────────
# Función librería (usada por la API)
# ─────────────────────────────────────────────

def run_extraction(input_folder, output_path, semana="?", bl_file=None, marchamo_file=None):
    """
    Ejecuta el pipeline completo y escribe el Excel en output_path.
    Retorna: {"rows": N, "warnings": [str, ...]}
    Lanza ValueError si falta algo indispensable (sin sys.exit, para poder
    usarse dentro de un servicio web).
    """
    if not os.path.isdir(input_folder):
        raise ValueError(f"carpeta no encontrada: {input_folder}")

    marchamo_lookup = load_marchamo_lookup(marchamo_file)

    # ── Localizar el PDF combinado de BLs ──
    if not bl_file:
        candidates = []
        for root, _dirs, files in os.walk(input_folder):
            for fname in files:
                if classify_file(fname) == 'BL':
                    candidates.append(os.path.join(root, fname))
        if candidates:
            bl_file = newest_file(candidates)

    if not bl_file or not os.path.isfile(bl_file):
        raise ValueError("no se encontró el PDF combinado de BLs (pásalo explícito o colócalo en la carpeta)")

    pages_by_code = split_bl_pdf(bl_file)

    bl_by_container = {}  # CONTENEDOR -> (bl_data, warning)
    lead_warnings = []
    for code, text in pages_by_code.items():
        bl_data, bl_warning = extract_bl_from_text(text, code)
        cont = bl_data["CONTENEDOR"]
        if cont == "???":
            lead_warnings.append(f"No se pudo extraer el contenedor de {code}")
            continue
        bl_by_container[cont] = (bl_data, bl_warning)

    # ── Facturas / fitos / manifiestos (sin importar estructura de carpetas) ──
    facturas_raw, fitos_raw, manifiestos_raw = classify_and_extract_all(input_folder, exclude_path=bl_file)
    facturas    = dedup_by_key(facturas_raw, "factura")
    fitos       = dedup_by_key(fitos_raw, "number")
    manifiestos = dedup_by_key(manifiestos_raw, "number")

    # ── Universo de contenedores: BL ∪ facturas ∪ fitos ──
    all_containers = set(bl_by_container.keys())
    for f in facturas:
        all_containers.update(f.get("containers", []))
    for fi in fitos:
        all_containers.update(fi.get("containers", []))

    if not all_containers:
        raise ValueError("no se pudo determinar ningún contenedor a partir de los documentos")

    rows, all_warnings = [], list(lead_warnings)
    for cont in sorted(all_containers):
        bl_data, bl_warning = bl_by_container.get(cont, (None, None))
        try:
            row, warns = build_row(cont, bl_data, bl_warning, facturas, fitos, manifiestos)
            rows.append(row)
            all_warnings.extend(warns)
        except Exception as e:
            all_warnings.append(f"ERROR procesando {cont}: {e}")
            rows.append({"BL": "???", "PROVEEDOR": "???", "CONTENEDOR": cont,
                         "TIPO": "???", "DESTINO": "???", "FACTURA": "???",
                         "ITEM": "???", "FITO": "???", "MANIFIESTO": "???"})

    rows.sort(key=lambda r: (r.get("BL", "???"), r.get("CONTENEDOR", "")))

    generate_excel(rows, output_path, semana, marchamo_lookup)

    return {"rows": len(rows), "warnings": all_warnings}
